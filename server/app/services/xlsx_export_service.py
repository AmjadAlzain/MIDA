"""
XLSX Export Service.

Generates XLSX files for:
- MIDA Certificate details with items and balances
- Item balance sheets with import history per port
- Bulk balance sheet exports (all items in a certificate)

Uses openpyxl for XLSX generation with styled headers and merged cells.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.mida_certificate import (
    ImportPort,
    MidaCertificate,
    MidaCertificateItem,
    MidaImportRecord,
)

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="333333")
LABEL_FONT = Font(bold=True, size=10, color="666666")
VALUE_FONT = Font(size=10, color="333333")
TABLE_HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TABLE_HEADER_FONT = Font(bold=True, size=10, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
PORT_DISPLAY_NAMES = {
    "port_klang": "Port Klang",
    "klia": "KLIA",
    "bukit_kayu_hitam": "Bukit Kayu Hitam",
}


def _format_decimal(value: Optional[Decimal]) -> str:
    """Format a decimal value for display."""
    if value is None:
        return "0"
    return f"{float(value):,.3f}".rstrip("0").rstrip(".")


def _format_date(value: Optional[date]) -> str:
    """Format a date value for display."""
    if value is None:
        return "-"
    return value.strftime("%Y-%m-%d")


def _set_column_widths(ws, widths: list[int]) -> None:
    """Set column widths for a worksheet."""
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_header_row(
    ws, row: int, headers: list[str], start_col: int = 1
) -> None:
    """Write a styled table header row."""
    for col, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = TABLE_HEADER_FILL
        cell.font = TABLE_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_info_row(
    ws, row: int, label: str, value: Any, label_col: int = 1, value_col: int = 2
) -> int:
    """Write a label-value info row with styling. Returns next row."""
    label_cell = ws.cell(row=row, column=label_col, value=label)
    label_cell.font = LABEL_FONT
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    value_cell = ws.cell(row=row, column=value_col, value=value)
    value_cell.font = VALUE_FONT
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    return row + 1


def _write_certificate_header(
    ws, certificate: MidaCertificate, start_row: int = 1
) -> int:
    """
    Write certificate information header with merged cells and styling.
    Returns the next available row.
    """
    row = start_row
    
    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value="MIDA Certificate Details")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    # Certificate info rows
    row = _write_info_row(ws, row, "Certificate Number:", certificate.certificate_number)
    row = _write_info_row(ws, row, "Company Name:", certificate.company_name)
    row = _write_info_row(ws, row, "Model Number:", certificate.model_number)
    row = _write_info_row(ws, row, "Status:", certificate.status.upper())
    row = _write_info_row(
        ws, row, "Exemption Period:",
        f"{_format_date(certificate.exemption_start_date)} to {_format_date(certificate.exemption_end_date)}"
    )
    
    return row + 1


def _write_item_header(
    ws,
    item: MidaCertificateItem,
    certificate: MidaCertificate,
    start_row: int = 1,
    include_certificate: bool = True,
) -> int:
    """
    Write item information header with merged cells and styling.
    Returns the next available row.
    """
    row = start_row
    
    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=f"Balance Sheet - Item #{item.line_no}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    if include_certificate:
        # Certificate info
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cert_cell = ws.cell(row=row, column=1, value="Certificate Information")
        cert_cell.font = SUBTITLE_FONT
        row += 1
        
        row = _write_info_row(ws, row, "Certificate Number:", certificate.certificate_number)
        row = _write_info_row(ws, row, "Company Name:", certificate.company_name)
        row += 1
    
    # Item info
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    item_cell = ws.cell(row=row, column=1, value="Item Information")
    item_cell.font = SUBTITLE_FONT
    row += 1
    
    row = _write_info_row(ws, row, "Line #:", item.line_no)
    row = _write_info_row(ws, row, "HS Code:", item.hs_code)
    row = _write_info_row(ws, row, "Item Name:", item.item_name)
    row = _write_info_row(ws, row, "UOM:", item.uom)
    row += 1
    
    # Quantities
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    qty_cell = ws.cell(row=row, column=1, value="Quantity Summary")
    qty_cell.font = SUBTITLE_FONT
    row += 1
    
    row = _write_info_row(ws, row, "Total Approved:", _format_decimal(item.approved_quantity))
    row = _write_info_row(ws, row, "Total Remaining:", _format_decimal(item.remaining_quantity))
    row += 1
    
    # Port breakdown
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    port_cell = ws.cell(row=row, column=1, value="Port Allocation (Approved / Remaining)")
    port_cell.font = SUBTITLE_FONT
    row += 1
    
    row = _write_info_row(
        ws, row, "Port Klang:",
        f"{_format_decimal(item.port_klang_qty)} / {_format_decimal(item.remaining_port_klang)}"
    )
    row = _write_info_row(
        ws, row, "KLIA:",
        f"{_format_decimal(item.klia_qty)} / {_format_decimal(item.remaining_klia)}"
    )
    row = _write_info_row(
        ws, row, "Bukit Kayu Hitam:",
        f"{_format_decimal(item.bukit_kayu_hitam_qty)} / {_format_decimal(item.remaining_bukit_kayu_hitam)}"
    )
    
    return row + 1


def generate_certificate_xlsx(
    certificate: MidaCertificate,
) -> bytes:
    """
    Generate an XLSX file for a MIDA certificate with all items and their balances.
    
    Format:
    - Certificate header info at top (merged cells, styled)
    - Table with all items showing approved/remaining quantities per port
    
    Args:
        certificate: The MIDA certificate with items loaded
        
    Returns:
        XLSX file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"
    
    # Write certificate header
    row = _write_certificate_header(ws, certificate)
    row += 1
    
    # Write items table title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    items_title = ws.cell(row=row, column=1, value="Certificate Items")
    items_title.font = SUBTITLE_FONT
    row += 1
    
    # Write items table header
    headers = [
        "Line #", "HS Code", "Item Name", "UOM",
        "Approved Qty", "Remaining Qty",
        "PK Approved", "PK Remaining",
        "KLIA Approved", "KLIA Remaining",
        "BKH Approved", "BKH Remaining",
    ]
    _write_header_row(ws, row, headers)
    row += 1
    
    # Write items
    for item in certificate.items:
        ws.cell(row=row, column=1, value=item.line_no).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item.hs_code).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item.item_name).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item.uom).border = THIN_BORDER
        ws.cell(row=row, column=5, value=float(item.approved_quantity or 0)).border = THIN_BORDER
        ws.cell(row=row, column=6, value=float(item.remaining_quantity or 0)).border = THIN_BORDER
        ws.cell(row=row, column=7, value=float(item.port_klang_qty or 0)).border = THIN_BORDER
        ws.cell(row=row, column=8, value=float(item.remaining_port_klang or 0)).border = THIN_BORDER
        ws.cell(row=row, column=9, value=float(item.klia_qty or 0)).border = THIN_BORDER
        ws.cell(row=row, column=10, value=float(item.remaining_klia or 0)).border = THIN_BORDER
        ws.cell(row=row, column=11, value=float(item.bukit_kayu_hitam_qty or 0)).border = THIN_BORDER
        ws.cell(row=row, column=12, value=float(item.remaining_bukit_kayu_hitam or 0)).border = THIN_BORDER
        
        # Number formatting for quantity columns
        for col in range(5, 13):
            ws.cell(row=row, column=col).number_format = '#,##0.000'
        
        row += 1
    
    # Set column widths
    _set_column_widths(ws, [8, 15, 40, 10, 14, 14, 14, 14, 14, 14, 14, 14])
    
    # Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_item_balance_sheet_xlsx(
    item: MidaCertificateItem,
    certificate: MidaCertificate,
    port: Optional[str] = None,
    import_records: Optional[list] = None,
) -> bytes:
    """
    Generate an XLSX file for an item's balance sheet (import history).
    
    If port is None, generates a workbook with 3 sheets (one per port).
    If port is specified, generates a single sheet for that port.
    
    Format per sheet:
    - Item and certificate info header at top
    - Import history table below
    
    Args:
        item: The certificate item
        certificate: The parent certificate
        port: Optional port filter (None = all ports as separate sheets)
        import_records: List of import records (if None, uses item.import_records)
        
    Returns:
        XLSX file as bytes
    """
    # Use provided import_records or fall back to item.import_records
    all_records = import_records if import_records is not None else list(item.import_records or [])
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    ports_to_export = [port] if port else ["port_klang", "klia", "bukit_kayu_hitam"]
    
    for port_key in ports_to_export:
        # Filter records for this port
        records = [r for r in all_records if r.port == port_key]
        
        # Create sheet
        sheet_name = PORT_DISPLAY_NAMES.get(port_key, port_key)
        ws = wb.create_sheet(title=sheet_name)
        
        # Write item header
        row = _write_item_header(ws, item, certificate)
        row += 1
        
        # Write import history title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        history_title = ws.cell(
            row=row, column=1,
            value=f"Import History - {sheet_name}"
        )
        history_title.font = SUBTITLE_FONT
        row += 1
        
        # Write import table header
        headers = [
            "Date", "Invoice #", "Line", "Form Reg No",
            "Quantity", "Balance Before", "Balance After", "Remarks"
        ]
        _write_header_row(ws, row, headers)
        row += 1
        
        # Write import records
        if records:
            for record in sorted(records, key=lambda r: (r.import_date, r.created_at)):
                ws.cell(row=row, column=1, value=record.import_date).border = THIN_BORDER
                ws.cell(row=row, column=2, value=record.invoice_number).border = THIN_BORDER
                ws.cell(row=row, column=3, value=record.invoice_line or "-").border = THIN_BORDER
                ws.cell(row=row, column=4, value=record.declaration_form_reg_no or "-").border = THIN_BORDER
                ws.cell(row=row, column=5, value=float(record.quantity_imported)).border = THIN_BORDER
                ws.cell(row=row, column=6, value=float(record.balance_before)).border = THIN_BORDER
                ws.cell(row=row, column=7, value=float(record.balance_after)).border = THIN_BORDER
                ws.cell(row=row, column=8, value=record.remarks or "-").border = THIN_BORDER
                
                # Number formatting
                for col in [5, 6, 7]:
                    ws.cell(row=row, column=col).number_format = '#,##0.000'
                
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            no_data_cell = ws.cell(row=row, column=1, value="No import records for this port")
            no_data_cell.font = Font(italic=True, color="999999")
            no_data_cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        _set_column_widths(ws, [12, 18, 8, 18, 14, 14, 14, 30])
    
    # Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_all_items_balance_sheets_xlsx(
    certificate: MidaCertificate,
    port: str,
    import_records_by_item: Optional[dict] = None,
) -> bytes:
    """
    Generate an XLSX file with balance sheets for ALL items in a certificate,
    for a specific port. Each item gets its own sheet.
    
    Args:
        certificate: The certificate with items loaded
        port: The port to export (port_klang, klia, bukit_kayu_hitam)
        import_records_by_item: Dict mapping item_id to list of import records
        
    Returns:
        XLSX file as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    port_display = PORT_DISPLAY_NAMES.get(port, port)
    
    for item in certificate.items:
        # Get records for this item from the dict, or fall back to item.import_records
        item_records = []
        if import_records_by_item is not None:
            item_records = import_records_by_item.get(item.id, [])
        else:
            item_records = list(item.import_records or [])
        
        # Filter records for this port
        records = [r for r in item_records if r.port == port]
        
        # Create sheet with safe name (max 31 chars for Excel)
        sheet_name = f"Item {item.line_no} - {item.hs_code}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Write item header
        row = _write_item_header(ws, item, certificate, include_certificate=True)
        row += 1
        
        # Write import history title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        history_title = ws.cell(
            row=row, column=1,
            value=f"Import History - {port_display}"
        )
        history_title.font = SUBTITLE_FONT
        row += 1
        
        # Write import table header
        headers = [
            "Date", "Invoice #", "Line", "Form Reg No",
            "Quantity", "Balance Before", "Balance After", "Remarks"
        ]
        _write_header_row(ws, row, headers)
        row += 1
        
        # Write import records
        if records:
            for record in sorted(records, key=lambda r: (r.import_date, r.created_at)):
                ws.cell(row=row, column=1, value=record.import_date).border = THIN_BORDER
                ws.cell(row=row, column=2, value=record.invoice_number).border = THIN_BORDER
                ws.cell(row=row, column=3, value=record.invoice_line or "-").border = THIN_BORDER
                ws.cell(row=row, column=4, value=record.declaration_form_reg_no or "-").border = THIN_BORDER
                ws.cell(row=row, column=5, value=float(record.quantity_imported)).border = THIN_BORDER
                ws.cell(row=row, column=6, value=float(record.balance_before)).border = THIN_BORDER
                ws.cell(row=row, column=7, value=float(record.balance_after)).border = THIN_BORDER
                ws.cell(row=row, column=8, value=record.remarks or "-").border = THIN_BORDER
                
                # Number formatting
                for col in [5, 6, 7]:
                    ws.cell(row=row, column=col).number_format = '#,##0.000'
                
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            no_data_cell = ws.cell(row=row, column=1, value="No import records for this port")
            no_data_cell.font = Font(italic=True, color="999999")
            no_data_cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        _set_column_widths(ws, [12, 18, 8, 18, 14, 14, 14, 30])
    
    # Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
