"""
Extract HSCODE Master data from Excel file.

Reads all sheets from the HSCODE Master Excel file and extracts:
- MIDA PART NAME -> Part Name
- MIDA 8 DIGIT -> HSCODE
- Unit Measurement -> UOM

Outputs to a CSV file for database seeding.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Optional

import openpyxl


def normalize_column_name(name: str) -> str:
    """Normalize column name for case-insensitive matching."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def find_column_index(header_row: tuple, target_names: list[str]) -> Optional[int]:
    """
    Find column index by matching against possible column names.
    
    Args:
        header_row: Tuple of column headers
        target_names: List of possible column names (lowercase)
        
    Returns:
        Column index (0-based) or None if not found
    """
    for i, cell in enumerate(header_row):
        normalized = normalize_column_name(cell)
        for target in target_names:
            if target in normalized:
                return i
    return None


def extract_hscode_master(
    excel_path: Path,
    output_csv_path: Path,
) -> dict:
    """
    Extract HSCODE master data from all sheets in an Excel file.
    
    Args:
        excel_path: Path to the Excel file
        output_csv_path: Path to output CSV file
        
    Returns:
        Dictionary with extraction statistics
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Track unique entries (keep first occurrence)
    seen_entries: set[str] = set()  # (part_name, hscode) as key
    all_records: list[dict] = []
    
    # Column name variations to search for
    part_name_cols = ["mida part name", "part name", "partname"]
    hscode_cols = ["mida 8 digit", "8 digit", "hscode", "hs code"]
    uom_cols = ["unit mesurement", "unit measurement", "uom", "unit"]
    
    stats = {
        "sheets_processed": 0,
        "total_rows": 0,
        "unique_records": 0,
        "duplicates_skipped": 0,
        "skipped_missing_data": 0,
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        if len(rows) < 2:  # Need header + at least 1 data row
            continue
            
        header_row = rows[0]
        
        # Find column indices
        part_name_idx = find_column_index(header_row, part_name_cols)
        hscode_idx = find_column_index(header_row, hscode_cols)
        uom_idx = find_column_index(header_row, uom_cols)
        
        if part_name_idx is None or hscode_idx is None or uom_idx is None:
            print(f"Warning: Sheet '{sheet_name}' missing required columns, skipping")
            print(f"  Found: part_name={part_name_idx}, hscode={hscode_idx}, uom={uom_idx}")
            continue
        
        stats["sheets_processed"] += 1
        
        # Process data rows
        for row in rows[1:]:
            stats["total_rows"] += 1
            
            # Extract values
            part_name = row[part_name_idx] if part_name_idx < len(row) else None
            hscode = row[hscode_idx] if hscode_idx < len(row) else None
            uom = row[uom_idx] if uom_idx < len(row) else None
            
            # Clean values
            part_name = str(part_name).strip() if part_name else ""
            hscode = str(hscode).strip() if hscode else ""
            uom = str(uom).strip().upper() if uom else ""
            
            # Skip rows with missing essential data
            if not part_name or not hscode or hscode == "None":
                stats["skipped_missing_data"] += 1
                continue
            
            # Normalize HSCODE (remove dots, ensure 8 digits)
            hscode = hscode.replace(".", "").replace("-", "")
            
            # Skip formula remnants
            if hscode.startswith("=") or part_name.startswith("="):
                stats["skipped_missing_data"] += 1
                continue
            
            # Deduplicate by (part_name, hscode) - keep first occurrence
            key = f"{part_name.lower()}|{hscode}"
            if key in seen_entries:
                stats["duplicates_skipped"] += 1
                continue
            
            seen_entries.add(key)
            
            # Normalize UOM
            uom_normalized = normalize_uom(uom)
            
            all_records.append({
                "Part Name": part_name,
                "HSCODE": hscode,
                "UOM": uom_normalized,
            })
    
    stats["unique_records"] = len(all_records)
    
    # Write to CSV
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Part Name", "HSCODE", "UOM"])
        writer.writeheader()
        writer.writerows(all_records)
    
    print(f"\nExtraction complete!")
    print(f"  Sheets processed: {stats['sheets_processed']}")
    print(f"  Total rows: {stats['total_rows']}")
    print(f"  Unique records: {stats['unique_records']}")
    print(f"  Duplicates skipped: {stats['duplicates_skipped']}")
    print(f"  Skipped (missing data): {stats['skipped_missing_data']}")
    print(f"\nOutput written to: {output_csv_path}")
    
    return stats


def normalize_uom(uom: str) -> str:
    """
    Normalize UOM value to standard format.
    
    Args:
        uom: Raw UOM value
        
    Returns:
        Normalized UOM: "KGM" or "UNIT"
    """
    if not uom:
        return "UNIT"
    
    uom_upper = uom.upper().strip()
    
    # KGM variations
    if uom_upper in ("KGM", "KG", "KGS", "KILOGRAM", "KILOGRAMS"):
        return "KGM"
    
    # UNIT variations
    if uom_upper in ("UNIT", "UNT", "UNITS", "PCS", "PC", "PIECE", "PIECES", "EA", "EACH", "NOS", "NO"):
        return "UNIT"
    
    # Default to UNIT
    return "UNIT"


if __name__ == "__main__":
    # Paths
    project_root = Path(__file__).parent.parent.parent
    excel_path = project_root / "web" / "010725 HSCODEMASTER (1).xlsx"
    output_csv_path = project_root / "server" / "hscode_master.csv"
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        exit(1)
    
    extract_hscode_master(excel_path, output_csv_path)
